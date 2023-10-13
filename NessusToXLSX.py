import xml.etree.ElementTree as ET
from openpyxl import Workbook
from tqdm import tqdm
import argparse
import os
from googletrans import Translator
wb = Workbook()
translator = Translator()
ws = wb.active
def VulnerabilityIsExsist(vulnerability, unique_list):
    temp = ReportHostItem()
    for vuln in unique_list:
        temp = vuln
        if vuln.plugin_name != vulnerability.plugin_name:
            continue
        if vuln.description != vulnerability.description:
            continue
        if vuln.severity != vulnerability.severity:
            continue
        return temp
    return None

def VulnerabilityBasedReport(report):
    unique_list = []
    # traverse for all elements
    for host in report:
        for vulnerability in host.reportHostItems:
            vuln = VulnerabilityIsExsist(vulnerability, unique_list) 
            if vuln is None or len(unique_list) == 0:
                vuln = Vulnerability(
                    plugin_name=vulnerability.plugin_name,
                    plugin_output=vulnerability.plugin_output,
                    solution=vulnerability.solution,
                    description=vulnerability.description,
                    cve=vulnerability.cve,
                    severity=vulnerability.severity,
                    severityID=vulnerability.severityID,
                    ipList=[]
                    )
                vuln.appendIp(host.ip)
                unique_list.append(vuln)
            else:
                vuln.appendIp(host.ip)
    return unique_list


class ReportHostItem:
    def __init__(self, plugin_name=None, plugin_output=None, solution = None,description=None,cve=None,severity=None,severityID=None):
        self.plugin_name = plugin_name
        self.plugin_output = plugin_output
        self.solution = solution
        self.description = description
        self.cve = cve
        self.severity = severity
        self.severityID = severityID
    def print(self):
        print(f"{self.plugin_name}")
class ReportHost:
    def __init__(self, ip, reportHostItems ):
        self.ip = ip
        self.reportHostItems = reportHostItems
    def print(self):
        print(f"{self.ip}")
        for item in self.reportHostItems:
            item.print()
class Vulnerability:
    def __init__(self, plugin_name, plugin_output, solution,description,cve,severity,severityID,ipList):
        self.ipList = ipList
        self.plugin_name = plugin_name
        self.plugin_output = plugin_output
        self.solution = solution
        self.description = description
        self.cve = cve
        self.severity = severity
        self.severityID = severityID
    def appendIp(self,ipAddress):
        self.ipList.append(ipAddress)
    def print(self):
        print(f"{self.plugin_name}")
      

def ParseReportHostItem(reportHostItem):
    severityID = reportHostItem.get("severity")
    plugin_name = ""
    plugin_output = ""
    description = ""
    solution = ""
    cve = ""
    for Item in reportHostItem:
        if Item.tag == "plugin_name":
            plugin_name = Item.text
        if Item.tag == "plugin_output":
            plugin_output = Item.text
        if Item.tag == "description":
            description = Item.text
        if Item.tag == "solution":
            solution = Item.text
        if Item.tag == "cve":
            if cve == "":
                cve = Item.text
            else:
                cve += ", " + Item.text
    if plugin_name =="" and plugin_output=="" and description == "" and cve =="":
        return None
    return ReportHostItem(
        plugin_name=plugin_name,
        plugin_output=plugin_output,
        solution=solution,
        description=description,
        cve=cve,
        severity=GetSeverityString(severityID),
        severityID=int(severityID)
        )

def ParseReportHost(reportHost):
    if reportHost.tag != "ReportHost":
        return
    ip = reportHost.get("name")
    reportItemArray = []
    for ReportItem in reportHost.findall("ReportItem"):
        item = ParseReportHostItem(ReportItem) 
        if item == None:
            continue
        reportItemArray.append(item)
    if len(reportItemArray) == 0:
        return
    return ReportHost(ip,reportItemArray)
def ParseReport(xml_content):
    root = ET.fromstring(xml_content).find("Report")
    reportHostArray = []
    for reportHost in root.findall("ReportHost"):
        reportHost = ParseReportHost(reportHost)
        if reportHost == None:
                continue 
        reportHostArray.append(reportHost) 
    return reportHostArray

    
def GetSeverityString(severity):
    if severity == "0":
        return "Informational"
    elif severity == "1":
        return "Low"
    elif severity == "2":
        return "Medium"
    elif severity == "3":
        return "High"
    elif severity == "4":
        return "Critical"
def check_severity(value):
    ivalue = int(value)
    if ivalue < 0 or ivalue > 4:
        raise argparse.ArgumentTypeError(f"Severity must be between 0 and 4, got {ivalue}")
    return ivalue
def check_input_file(value):
    if not os.path.exists(value):
        raise argparse.ArgumentTypeError(f"Input file '{value}' does not exist")
    return value
def main(input_file, output_file, severity, cve,IpBased,translate):
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    print(f"Severity: {severity}")
    print(f"CVE flag: {cve}")

    with open(input_file, 'r', encoding='utf-8') as file:
        xml_content = file.read()
        report = ParseReport(xml_content)
        row = 2
        if IpBased == False:
            vreport = VulnerabilityBasedReport(report)
            for vuln in tqdm(vreport, "Progress: "):
                if vuln.severityID < severity:
                    continue
                if cve and vuln.cve == "":
                    continue
                ws.cell(row=row, column=1, value=vuln.plugin_name)
                ws.cell(row=row, column=1, value=translator.translate(vuln.plugin_name, dest=translate).text)
                ws.cell(row=row, column=2, value=vuln.description)
                ws.cell(row=row, column=3, value=translator.translate(vuln.description, dest=translate).text)
                ws.cell(row=row, column=4, value=vuln.solution)
                ws.cell(row=row, column=5, value=translator.translate(vuln.solution, dest=translate).text)
                ws.cell(row=row, column=6, value=vuln.cve)
                ws.cell(row=row, column=7, value=translator.translate(vuln.severity, dest=translate).text)
                ws.cell(row=row, column=8, value=', '.join(vuln.ipList))
                ws.cell(row=row, column=9, value=vuln.plugin_output)
                row = row + 1


        else:
            for reportHosts in tqdm(report,"Host Porgress: "):
                for hostItem in tqdm(reportHosts.reportHostItems,"Host Vulns: "):
                    if hostItem.severityID < severity:
                        continue
                    if cve and hostItem.cve == "":
                        continue
                    ws.cell(row=row, column=1, value=reportHosts.ip)
                    ws.cell(row=row, column=2, value=hostItem.plugin_name)
                    ws.cell(row=row, column=2, value=translator.translate(hostItem.plugin_name, dest=translate).text)
                    ws.cell(row=row, column=3, value=hostItem.description)
                    ws.cell(row=row, column=4, value=translator.translate(hostItem.description, dest=translate).text)
                    ws.cell(row=row, column=5, value=hostItem.solution)
                    ws.cell(row=row, column=6, value=translator.translate(hostItem.solution, dest=translate).text)
                    ws.cell(row=row, column=8, value=translator.translate(hostItem.severity, dest=translate).text)
                    ws.cell(row=row, column=7, value=hostItem.cve)
                    ws.cell(row=row, column=9, value=hostItem.plugin_output)
                    row = row + 1
    
    wb.save(output_file)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Enjoy brother !!")
    parser.add_argument("--input","-i", dest="input_file", type=check_input_file, default="scan.nessus", help="Path to the input file")
    parser.add_argument("--output","-o", dest="output_file", default="output.xlsx", help="Path to the output file")
    parser.add_argument("--severity","-s", dest="severity", type=check_severity, default=2, choices=[0, 1, 2, 3, 4], help="(Default 2) Minimum severity level: 0 (informational), 1 (low), 2 (medium), 3 (high), 4 (critical)")
    parser.add_argument("--cve","-c", action="store_true", help="Just Include if has CVE information")
    parser.add_argument("--translate","-t",default="tr", help="Just Include if has CVE information")
    parser.add_argument("-IpBased","-ib",action="store_true", help="IP based output")
    args = parser.parse_args()

    main(args.input_file, args.output_file, args.severity, args.cve,args.IpBased,args.translate)

